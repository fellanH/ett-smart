#!/usr/bin/env python3
"""
Script to convert blue-collar-companies.csv to a formatted Excel file.
"""

import csv

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def format_excel_file(excel_path):
    """Apply formatting to the Excel file."""
    wb = load_workbook(excel_path)
    ws = wb.active
    
    # Define styles
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Format header row
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = border
    
    # Format data rows
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(vertical='top', wrap_text=True)
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        
        # Set width with some padding, but cap at 50
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Freeze header row
    ws.freeze_panes = 'A2'
    
    # Set row height for header
    ws.row_dimensions[1].height = 30
    
    wb.save(excel_path)
    print(f"✓ Formatted Excel file saved: {excel_path}")

def main():
    csv_file = 'blue-collar-companies.csv'
    excel_file = 'blue-collar-companies.xlsx'
    
    print(f"Reading CSV file: {csv_file}")
    # Read CSV with proper encoding for Swedish characters.
    # This CSV may contain rows with too many/few commas; normalize each row to header length.
    with open(csv_file, "r", encoding="utf-8", newline="") as f:
        reader = csv.reader(f, delimiter=",")
        header = next(reader)
        expected = len(header)

        rows = []
        for i, row in enumerate(reader, start=2):  # 1-based file line numbers
            if len(row) > expected:
                # Merge extra fields into the last column to preserve data.
                row = row[: expected - 1] + [",".join(row[expected - 1 :])]
            elif len(row) < expected:
                row = row + [""] * (expected - len(row))
            rows.append(row)

    df = pd.DataFrame(rows, columns=header)
    
    print(f"Found {len(df)} rows and {len(df.columns)} columns")
    print(f"Columns: {', '.join(df.columns.tolist())}")
    
    print(f"Creating Excel file: {excel_file}")
    # Write to Excel
    df.to_excel(excel_file, index=False, engine='openpyxl')
    
    print("Applying formatting...")
    format_excel_file(excel_file)
    
    print(f"\n✓ Successfully created formatted Excel file: {excel_file}")

if __name__ == '__main__':
    main()
